"""
YaoWo Tech - E-Commerce Scraper
Supports: 1688, Shein
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading, os, sys, re, json, time, subprocess
from datetime import datetime
sys.path.insert(0,os.path.dirname(os.path.abspath(__file__)))
# === Global exception hook: log unhandled errors to file ===
_LOG_FILE=os.path.join(os.path.dirname(os.path.abspath(__file__)),'error.log')
def _excepthook(etype,value,tb):
    import traceback as _tb
    msg=''.join(_tb.format_exception(etype,value,tb))
    try:
        with open(_LOG_FILE,'a',encoding='utf-8') as _f:
            _f.write(f'[{datetime.now()}] {msg}\n')
    except:pass
    sys.__excepthook__(etype,value,tb)
sys.excepthook=_excepthook
threading.excepthook=lambda args:_excepthook(args.exc_type,args.exc_value,args.exc_traceback)
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__),".env"))
except:pass

# === Portable auto-detect paths ===
_ROOT=os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)),'..'))
_CHROME_PROFILE=os.path.join(_ROOT,'data','chrome_profile').replace('\\\\?\\','')
_CLOAK=os.path.join(_ROOT,'cloakbrowser','chrome.exe')
_CLOAKBROWSER=_CLOAK if os.path.exists(_CLOAK) else ''
_LOG_FILE=os.path.join(_ROOT,'scraper.log')
try:
    os.makedirs(_CHROME_PROFILE,exist_ok=True)
except:
    pass

TEMPLATE_HEADERS=["ParentSKU","SKU","产品标题","产品描述","产品标签","品牌","UPC","LandingPageUrl","MSRP",
    "颜色","尺寸值","来源url","价格","运费","数量","最小送达时间","最大送达时间",
    "主图","附图1","附图2","附图3","附图4","附图5","附图6","附图7","附图8","附图9","附图10",
    "附图11","附图12","附图13","附图14","附图15","附图16","附图17","附图18","附图19","附图20",
    "产品地址","描述图片","变种图（URL）地址","ShippingWeight","VariantMainImageURL","DescHtml",
    "PubField1","PubField2","PubField3","PubField4","PubField5","PubField6",
    "FenleiId","FenleiName","Category","VideoUrl","原标题","语种",
    "包裹长度","包裹宽度","包裹高度","EAN","预留1","预留2","预留3","产品品论数","Highlight","SizeUrl",
    "详情图1","详情图2","详情图3","详情图4","详情图5","详情图6","详情图7","详情图8","详情图9","详情图10",
    "详情图11","详情图12","详情图13","详情图14","详情图15","详情图16","详情图17","详情图18","详情图19","详情图20"]

def extract_json(text,key):
    idx=text.find(f'"{key}"')
    if idx<0:return None
    idx=text.find('{',idx)
    if idx<0:return None
    depth=1;i=idx+1
    while i<len(text)and depth>0:
        if text[i]=='{':depth+=1
        elif text[i]=='}':depth-=1
        i+=1
    return text[idx:i]

def extract_shein(url, pause_event=None, scraped_skus=None, captcha_cb=None, shared_page=None, check_skip=None):
    from Shein import Shein;from bs4 import BeautifulSoup
    if scraped_skus is None:scraped_skus=set()
    def _skip():return check_skip()if check_skip else False
    def so(u):
        # Quick SKU pre-check using shared page (navigate, don't create/close)
        if scraped_skus:
            for _retry in range(2):
                try:
                    if shared_page:
                        shared_page.goto(u,timeout=30000,wait_until="domcontentloaded")
                        shared_page.evaluate("window.blur()")
                        time.sleep(0.5)
                        _html=shared_page.content()
                        _sku_m=re.search(r'SKU:\s*([\w\d]+)',_html)
                        if _sku_m and _sku_m.group(1) in scraped_skus:
                            return None
                        break
                    else:
                        from patchright.sync_api import sync_playwright as _sp
                        _pw=_sp().start()
                        _ctx=_pw.chromium.launch_persistent_context(
                            user_data_dir=_CHROME_PROFILE,
                            executable_path=_CLOAKBROWSER if _CLOAKBROWSER else None,headless=False,
                            args=["--no-sandbox"],
                            ignore_default_args=["--enable-automation","--enable-unsafe-swiftshader"],
                            viewport={"width":1920,"height":1080})
                        _pg=_ctx.new_page()
                        _pg.goto(u,timeout=30000,wait_until="domcontentloaded")
                        _pg.evaluate("window.blur()")
                        time.sleep(2)
                        _html=_pg.content()
                        _pg.close();_ctx.close();_pw.stop()
                        _sku_m=re.search(r'SKU:\s*([\w\d]+)',_html)
                        if _sku_m and _sku_m.group(1) in scraped_skus:
                            return None
                        break
                except Exception as _e:
                    if _retry==0:
                        time.sleep(2)
                        continue
                    break  # failed twice, proceed with full scrape anyway

        if pause_event:pause_event.wait()
        if _skip():return None
        s=Shein(u,external_page=shared_page);s.set_captcha_callback(captcha_cb)
        if pause_event:s.set_pause_event(pause_event)
        if check_skip:s.set_skip_check(check_skip)
        d=s.scrape()
        if not d:return None
        pd=os.path.join(os.path.dirname(__file__),"Outputs",d.get("product_name_safe","Unknown"))
        pp=os.path.join(pd,"page.html")
        if not os.path.exists(pp):return None
        with open(pp,'r',encoding='utf-8')as f:html=f.read()
        soup=BeautifulSoup(html,"html.parser")
        # Product-level SKU from HTML
        prod_sku=""
        sku_el=soup.find('span',string=re.compile(r'SKU:\s*[\w\d]+'))
        if not sku_el:
            sku_el=soup.select_one('.product-intro__head-sku-text')
        if sku_el:
            sku_text=sku_el.text.strip()
            m_sku=re.search(r'SKU:\s*([\w\d]+)',sku_text)
            if m_sku:prod_sku=m_sku.group(1)
        # Double-check SKU (in case pre-check was skipped)
        if prod_sku and scraped_skus and prod_sku in scraped_skus:
            return None
        # skcImages: accurate per-color images from HTML
        skc_imgs=[]
        sim=re.search(r'"skcImages"\s*:\s*\[(.*?)\]',html,re.DOTALL)
        if sim:
            raw_imgs=re.findall(r'"([^"]*ltwebstatic[^"]*)"',sim.group(1))
            for u in raw_imgs:
                u='https:'+u if u.startswith('//')else u
                u=re.sub(r'_thumbnail_\d+x\d+','_thumbnail_900x',u)
                skc_imgs.append(u)
        # JSON-LD: all colors (each item = one color)
        vi=[]
        m=re.search(r'<script[^>]*id="goodsDetailSchema"[^>]*>(.*?)</script>',html,re.DOTALL)
        color_entries=[]
        if m:
            try:g=json.loads(m.group(1))
            except:g=None
            if g:
                for it in(g if isinstance(g,list)else[g]):
                    try:
                        if not isinstance(it,dict)or'hasVariant'not in it:continue
                        cn=it.get('color','')or'Default'
                        sizes=[];szpr={}
                        for v in it['hasVariant']:
                            n=v.get('name','')
                            sz=re.search(r'(?:Size|사이즈)\s+([\w\d]+)',n,re.I)
                            if not sz:sz=re.search(r'(?:Size\s+)?(EUR\d+|US\d+|UK\d+)',n,re.I)
                            s2=sz.group(1)if sz else''
                            if s2:
                                if s2 not in sizes:sizes.append(s2)
                                o=v.get('offers',{});vp=o.get('price','')if isinstance(o,dict)else''
                                if vp:szpr[s2]=vp
                                v2=v.get('image',[])
                                u=(re.sub(r'_thumbnail_\d+x\d+','_thumbnail_900x',v2[0])if isinstance(v2,list)and v2 else
                                   re.sub(r'_thumbnail_\d+x\d+','_thumbnail_900x',v2)if isinstance(v2,str)and v2 else'')
                                if u:vi.append(u)
                        # Filter out non-size text (size chart refs, Korean words)
                        _size_blacklist = ['차트','차드','표를','표','사이즈','가이드','크기','선택','선택하세요']
                        sizes=[s for s in sizes if s not in _size_blacklist]
                        try:sizes.sort(key=lambda s:(re.match(r'([A-Z]*)(\d+)',s)and(re.match(r'([A-Z]*)(\d+)',s).group(1),int(re.match(r'([A-Z]*)(\d+)',s).group(2))))or(s,0))
                        except:pass
                        if not sizes:sizes=["One Size"]
                        # Merge skcImages + vi, dedup by stem
                        seen=set();ui=[]
                        for u in skc_imgs+vi:
                            if not u.startswith('http'):continue
                            fn=u.split('/')[-1].split('?')[0].lower()
                            if fn.startswith('visa')or fn.startswith('mastercard'):continue
                            stem=re.sub(r'_thumbnail_\d+x\d+','',fn)
                            stem=re.sub(r'\.(jpg|jpeg|png|webp|gif)$','',stem)
                            if stem not in seen:seen.add(stem);ui.append(u)
                        pr=szpr.get(sizes[0],'')if sizes and szpr else''
                        if not pr:
                            for v in it['hasVariant'][:1]:
                                o=v.get('offers',{})
                                if isinstance(o,dict)and o.get('price'):pr=o.get('price','')
                        if not pr:
                            pr=f"{d.get('current_price_integer','')}.{d.get('current_price_decimal','00')}"if d.get('current_price_integer')else""
                        ci_img=vi[0]if vi else(skc_imgs[0]if skc_imgs else'')
                        color_entries.append({"color":cn,"sizes":sizes,"images":ui,"price":pr,
                            "size_prices":szpr,"color_image":ci_img,"sku":prod_sku})
                    except:continue
        if not color_entries:
            color_entries=[{"color":"Default","sizes":["One Size"],"images":[],"price":"","size_prices":{},"color_image":"","sku":""}]
        # Extract Shein category from meta tags
        shein_cat=""
        for meta in soup.find_all("meta"):
            nm=(meta.get("name","")+meta.get("property","")).lower()
            if"category"in nm:
                shein_cat=meta.get("content","")
                break
        return{"name":d.get("name",""),"description":d.get("description",""),"overview":d.get("overview","") or "","html":html,
               "discount":d.get("discount_percentage",""),"colors":color_entries,
               "shein_category":shein_cat}
    r=so(url)
    if r is None:return None  # SKU duplicate
    if not r:return[]
    sid=re.search(r'-p-(\d+)',url);ps=sid.group(1)if sid else"unknown"
    all_colors=list(r["colors"])
    seen_colors=set(c["color"]for c in all_colors)
    main_gid=re.search(r'-p-(\d+)',url).group(1)if re.search(r'-p-(\d+)',url)else""
    main_sku=r["colors"][0].get("sku","")if r["colors"]else""
    fallback_sizes=list(r["colors"][0]["sizes"]) if r["colors"] else["One Size"]
    fallback_price=r["colors"][0]["price"] if r["colors"] else""
    # Extract additional colors from HTML swatch data
    color_blocks=re.findall(r'"attr_name":"색","attr_value_id":"(\d+)","attr_value":"([^"]+)"[^}]*?"goods_id":"(\d+)"[^}]*?"goods_color_image":"([^"]+)"[^}]*?"goods_image":"([^"]+)"',r["html"])
    if not color_blocks:
        color_blocks=re.findall(r'"attr_value":"([^"]+)","goods_id":"(\d+)"[^}]*?"goods_color_image":"([^"]+)"[^}]*?"goods_image":"([^"]+)"',r["html"])
    for cb in color_blocks:
        if len(cb)==5:_,attr_val,gid,color_img,goods_img=cb
        elif len(cb)==4:attr_val,gid,color_img,goods_img=cb
        else:continue
        if gid==main_gid:continue
        label=f"{attr_val}"
        # Don't skip duplicates — collect all
        seen_colors.add(label)
        # Scrape this color's page for sizes/prices/images
        if pause_event:pause_event.wait()
        if _skip():continue
        bu=re.sub(r'\?.*','',url).rstrip('/')
        cu=re.sub(r'-p-\d+\.html',f'-p-{gid}.html',bu)
        extra_sizes=[];extra_price='';extra_imgs=[];extra_sku=main_sku
        try:
            cr=so(cu)
            if cr and cr["colors"]:
                ce=cr["colors"][0]
                extra_sizes=ce.get("sizes",[])
                extra_price=ce.get("price","")
                extra_imgs=ce.get("images",[])
                extra_sku=ce.get("sku","")or main_sku
        except:pass
        _size_bl=['차트','차드','표를','표','사이즈','가이드','크기','선택','선택하세요']
        extra_sizes=[s for s in extra_sizes if s not in _size_bl]
        if not extra_sizes:extra_sizes=list(fallback_sizes)
        if not extra_price:extra_price=fallback_price
        if not extra_imgs:
            swatch_img=f"https:{goods_img}"if goods_img.startswith('//')else goods_img
            if swatch_img:extra_imgs=[swatch_img]
        ci_img=extra_imgs[0]if extra_imgs else(f"https:{color_img}"if color_img.startswith('//')else color_img)
        all_colors.append({"color":label,"sizes":extra_sizes,"images":extra_imgs,
            "price":extra_price,"size_prices":{},"color_image":ci_img,"sku":extra_sku})
    # Split into ParentSKU groups: max 3 unique colors per group, no same-name per group
    from collections import Counter
    # Count color frequencies and sort by frequency (most first)
    color_names=[c["color"]for c in all_colors]
    freq=Counter(color_names)
    groups=[]  # each group = list of color entries, max 3, unique names
    remaining=list(all_colors)
    while remaining:
        group=[]
        used_names=set()
        new_remaining=[]
        for entry in remaining:
            cn=entry["color"]
            if cn not in used_names and len(group)<3:
                group.append(entry);used_names.add(cn)
            else:
                new_remaining.append(entry)
        remaining=new_remaining
        if group:
            groups.append(group)
    # Assign sequence suffix to ParentSKU
    ps_base=ps
    results=[]
    for gi,group in enumerate(groups):
        ps_group=f"{ps_base}_{gi+1}"if len(groups)>1 else ps_base
        for c in group:
            results.append({"parent_sku":ps_group,"name":r["name"],"url":url,"price":c["price"],"discount":r["discount"],
                "description":r["description"],"color":c["color"],"sizes":c["sizes"],"images":c["images"],
                "size_prices":c.get("size_prices",{}),"color_images":{c["color"]:c.get("color_image","")},
                "shein_category":r.get("shein_category",""),"sku":c.get("sku","")})
    return results

def extract_1688(url, shared_page=None, pause_event=None):
    from bs4 import BeautifulSoup
    from patchright.sync_api import sync_playwright
    if pause_event:pause_event.wait()
    try:
        if shared_page:
            shared_page.goto(url,timeout=60000,wait_until="domcontentloaded")
            shared_page.evaluate("window.blur()")
            time.sleep(0.5);html=shared_page.content()
        else:
            pw=sync_playwright().start()
            ctx=pw.chromium.launch_persistent_context(
                user_data_dir=_CHROME_PROFILE,
                executable_path=_CLOAKBROWSER if _CLOAKBROWSER else None,headless=False,
                args=["--no-sandbox","--disable-features=TranslateUI"],
                ignore_default_args=["--enable-automation","--enable-unsafe-swiftshader"],
                viewport={"width":1920,"height":1080})
            page=ctx.new_page();page.goto(url,timeout=60000,wait_until="domcontentloaded")
            time.sleep(2);html=page.content();page.close();ctx.close();pw.stop()
        soup=BeautifulSoup(html,"html.parser");base="https:"if url.startswith("https")else"http:"
        title="";t=soup.select_one('div.module-od-title h1')or soup.select_one('h1')
        if t:title=t.text.strip()
        if not title and soup.title:title=soup.title.text.strip()
        # --- Product images: gallery img tags + data-src with ibank path ---
        images=[]
        for sel in['img.preview-img','.od-gallery-img img','.module-od-picture-gallery img','[data-src]']:
            for img in soup.select(sel):
                src=img.get('src')or img.get('data-src')or''
                if not src:continue
                if src.startswith('//'):src='https:'+src
                elif not src.startswith('http'):src=base+src
                if'cbu01.alicdn.com/img/ibank/'not in src:continue
                # Normalize URL: strip _.webp suffix, restore .jpg
                src=re.sub(r'\.jpg_\.webp$','.jpg',src)
                if src not in images:images.append(src)
        # --- Rest: colors, sizes, prices from DOM ---
        colors,sizes,vi,skus,ci=[],[],[],[],{}
        sm_json=extract_json(html,'skuModel')
        if sm_json:
            try:
                sm=json.loads(sm_json)
                for prop in sm.get('skuProps',[]):
                    for v in prop.get('value',[]):
                        if prop.get('prop','')in('颜色','Color','color'):
                            cn=v.get('name','');img=v.get('imageUrl','')
                            if cn not in colors:colors.append(cn)
                            if img:
                                img=img.replace('_sum.jpg','.jpg').replace('_.webp','.jpg')
                                if img not in vi:vi.append(img)
                                if cn not in ci:ci[cn]=img
                        else:
                            sz=v.get('name','')
                            if sz and sz not in sizes:sizes.append(sz)
                for key,entry in sm.get('skuInfoMap',{}).items():
                    attrs=entry.get('specAttrs','');parts=attrs.split('&gt;')
                    color=parts[0]if len(parts)>0 else'';size=parts[1]if len(parts)>1 else''
                    stock=entry.get('canBookCount',0)or entry.get('saleCount',0)
                    if stock<=0:continue
                    if size and size not in sizes:sizes.append(size)
                    skus.append({'color':color,'size':size,'stock':stock,'skuId':entry.get('skuId','')})
            except:pass
        if not sizes:
            for item in soup.select('div.module-od-sku-selection .expand-view-list>.expand-view-item'):
                spec=item.select_one('.item-label')
                if spec:sz=spec.text.strip()
                if sz and sz not in sizes:sizes.append(sz)
        if not sizes:sizes=["One Size"]
        spr_m=re.search(r'"skuRangePrices"\s*:\s*\[([^\]]+)\]',html);price=""
        if spr_m:
            try:
                tiers=json.loads('['+spr_m.group(1)+']')
                if tiers:price=tiers[0].get('price','')
            except:pass
        shop="";st=soup.find('a',class_='shop-company-name')or soup.select_one('.shop-name')
        if st:shop=(st.h1.get('title')or st.h1.text.strip())if st.h1 else st.text.strip()
        def sk(s):m=re.match(r'([A-Z]*)(\d+)',s);return(m.group(1),int(m.group(2)))if m else(s,0)
        try:sizes.sort(key=sk)
        except:pass
        sku_match=re.search(r'/offer/(\d+)\.html',url)
        if not sku_match:sku_match=re.search(r'(\d{8,})',url)
        ps=sku_match.group(1)if sku_match else"unknown"
        if colors:
            return[{"parent_sku":ps,"name":title,"url":url,"price":price,"discount":"","description":shop,
                    "color":c,"sizes":list(set(s['size']for s in skus if s['color']==c))or sizes,
                    "images":images[:20],"variant_images":vi,"shop":shop,"skus":[],
                    "color_images":ci}for c in colors]
        return[{"parent_sku":ps,"name":title,"url":url,"price":price,"discount":"","description":shop,
                "color":"","sizes":sizes,"images":images[:20],"variant_images":vi,"shop":shop,"skus":[],
                "color_images":ci}]
    except Exception as e:
        print(f"[1688 ERROR] {e}");import traceback;traceback.print_exc();return[]

def extract_aliexpress(url, shared_page=None, pause_event=None):
    from AliExpress import AliExpress
    if pause_event:pause_event.wait()
    try:
        ae=AliExpress(url,external_page=shared_page,skip_media=True)
        d=ae.scrape()
        if not d:return[]

        title=d.get("name","")
        # Use OLD price (original), not current sale price
        old_int=d.get('old_price_integer','')
        old_dec=d.get('old_price_decimal','00')
        price=f"{old_int}.{old_dec}"if old_int and old_int!='N/A'else""
        # Current/sale price as secondary
        cur_int=d.get('current_price_integer','')
        cur_dec=d.get('current_price_decimal','00')
        cur_price=f"{cur_int}.{cur_dec}"if cur_int else""
        # Use all gallery images (not just main_image)
        images=d.get("images",[])or[]
        variant_images=d.get("variant_images",[])or[]
        sizes=d.get("sizes",[])or["One Size"]
        tags=d.get("tags",[])or[]
        product_tag=d.get("product_tag","")
        color_images=d.get("color_images",{})or{}

        # SKU / parent ID
        sku_match=re.search(r'/item/(\d+)\.html',url)
        if not sku_match:sku_match=re.search(r'(\d{10,})',url)
        ps=sku_match.group(1)if sku_match else"unknown"

        # Multi-SKU: one entry per color variant
        variants=d.get("variants",[])or[]
        desc_images=d.get("description_images",[])or[]
        if variants:
            result=[]
            for v in variants:
                v_old=v.get("old_price")or price
                vsizes=v.get("sizes")or sizes
                # Per-variant color image
                v_color_img = v.get("color_image","")
                v_color_dict = {v.get("color",""): v_color_img} if v_color_img else color_images
                result.append({
                    "parent_sku":ps,"name":title,"url":url,"sku":f"{ps}_{v.get('color','')}",
                    "price":v_old,"cur_price":cur_price,"discount":d.get("discount_percentage",""),
                    "description":d.get("description",""),"overview":d.get("overview","") or "",
                    "color":v.get("color",""),
                    "sizes":vsizes,"images":images[:20],
                    "variant_images":variant_images,
                    "shop":"AliExpress","skus":[],
                    "color_images":v_color_dict,
                    "color_image":v_color_img,
                    "tags":tags,
                    "product_tag":product_tag,
                    "description_images":desc_images,
                    "sold_out":v.get("sold_out",False),
                })
            return result

        # No variants: single entry
        return[{
            "parent_sku":ps,"name":title,"url":url,"sku":ps,"price":price,"cur_price":cur_price,
            "discount":d.get("discount_percentage",""),
            "description":d.get("description",""),"overview":d.get("overview","") or "","color":"",
            "sizes":sizes,"images":images[:20],
            "variant_images":variant_images,"shop":"AliExpress","skus":[],
            "color_images":color_images,
            "tags":tags,
            "product_tag":product_tag,
            "description_images":desc_images,
        }]
    except Exception as e:
        print(f"[ALIEXPRESS ERROR] {e}");import traceback;traceback.print_exc();return[]

class App:
    BG="#E3F2FD";P="#1565C0";A="#0D47A1";G="#00897B";O="#EF6C00";T="#263238";W="#FFFFFF"
    UF=os.path.join(os.path.dirname(os.path.abspath(__file__)),"saved_urls.txt")
    SKU_FILE=os.path.join(os.path.dirname(os.path.abspath(__file__)),"scraped_skus.txt")
    SESSION_FILE=os.path.join(os.path.dirname(os.path.abspath(__file__)),".scraper_session.json")

    def __init__(self):
        self.rt=tk.Tk();self.rt.title("耀我科技 - 电商采集器");self.rt.geometry("920x720");self.rt.configure(bg=self.BG)
        # Manual browser mode (init early, before UI)
        self._manual_pw=None;self._manual_ctx=None;self._manual_page=None
        self._manual_status=tk.StringVar(value="浏览器未启动")
        # Restore last window size
        try:
            if os.path.exists(self.SESSION_FILE):
                with open(self.SESSION_FILE,'r',encoding='utf-8')as f:
                    geo=json.load(f).get('_win_geo','')
                if geo:self.rt.geometry(geo)
        except:pass
        top=tk.Frame(self.rt,bg=self.BG);top.pack(fill=tk.X,padx=20,pady=(10,3))
        tk.Label(top,text="耀我科技",font=("Microsoft YaHei",20,"bold"),bg=self.BG,fg=self.A).pack(side=tk.LEFT)
        tk.Label(top,text="· 电商采集器",font=("Arial",12),bg=self.BG,fg=self.P).pack(side=tk.LEFT,padx=3)
        tk.Label(top,text="    平台:",font=("Arial",11,"bold"),bg=self.BG,fg=self.T).pack(side=tk.LEFT,padx=(20,0))
        self.pl=tk.StringVar(value="1688")
        for v,t in[("shein","Shein"),("1688","1688"),("aliexpress","AliExpress")]:
            tk.Radiobutton(top,text=t,variable=self.pl,value=v,font=("Arial",11),bg=self.BG,activebackground=self.BG,selectcolor=self.BG).pack(side=tk.LEFT,padx=5)
        tk.Label(top,text="  Shein站点:",font=("Arial",9),bg=self.BG,fg=self.T).pack(side=tk.LEFT,padx=(15,0))
        self.sr=tk.StringVar(value="US")
        for v,t in[("US","US"),("KR","KR")]:
            tk.Radiobutton(top,text=t,variable=self.sr,value=v,font=("Arial",9),bg=self.BG,activebackground=self.BG,selectcolor=self.BG).pack(side=tk.LEFT,padx=2)
        tk.Frame(self.rt,height=1,bg="#BBDEFB").pack(fill=tk.X,padx=20)

        # === Two-tab layout ===
        pw=tk.PanedWindow(self.rt,orient=tk.VERTICAL,bg=self.BG,bd=0,sashwidth=4,sashrelief="raised")
        pw.pack(fill=tk.BOTH,expand=True,padx=20,pady=(5,0))
        nb=ttk.Notebook(pw);pw.add(nb,stretch="always")

        # Tab 1: Scrape
        t1=tk.Frame(nb,bg=self.BG);nb.add(t1,text="  采集商品  ")
        tk.Label(t1,text="商品 URL (一行一个，关闭自动保存):",font=("Arial",10,"bold"),bg=self.BG,fg=self.T).pack(anchor=tk.W,padx=10,pady=(8,2))
        self.ut=scrolledtext.ScrolledText(t1,height=8,font=("Consolas",10),relief="solid",bd=1,bg=self.W)
        self.ut.pack(fill=tk.BOTH,expand=True,padx=10,pady=2)
        self.ut_label=tk.Label(t1,text="0 URLs",font=("Arial",8),bg=self.BG,fg="#78909C")
        self.ut_label.pack(anchor=tk.W,padx=10)
        if os.path.exists(self.UF):self.ut.insert("1.0",open(self.UF,'r',encoding='utf-8').read())
        self.ut.bind("<KeyRelease>",lambda e:self._su())
        self._update_url_count()
        self.rt.protocol("WM_DELETE_WINDOW",self._oc)
        uf=tk.Frame(t1,bg=self.BG);uf.pack(fill=tk.X,padx=10,pady=3)
        for t,c,cmd in[("删除选中行","#E53935",self._ds),("去重","#7B1FA2",self._dd),("清空URL","#78909C",self._cu),("全选复制","#FF6F00",self._ca)]:
            tk.Button(uf,text=t,font=("Arial",9),bg=c,fg=self.W,padx=8,pady=2,relief="flat",cursor="hand2",command=cmd).pack(side=tk.LEFT,padx=2)
        bf=tk.Frame(t1,bg=self.BG);bf.pack(fill=tk.X,padx=10,pady=5)
        self._btns={}
        for t,c,cmd,b in[("开始",self.G,self._start,True),("暂停","#E53935",self._pause,False),("继续","#1565C0",self._resume,False),("跳过此SKU","#F4511E",self._skip_sku,False),("结束","#78909C",self._end,False),("导出",self.O,self._ex,False),("登录",self.A,self._li,False),("清空","#78909C",self._cl,False),("恢复上次","#FF6F00",self._restore_session,False)]:
            btn=tk.Button(bf,text=t,font=("Arial",10,"bold")if b else("Arial",10),bg=c,fg=self.W,padx=16,pady=5,relief="flat",cursor="hand2",command=cmd)
            btn.pack(side=tk.LEFT,padx=4)
            self._btns[t]=btn
            if t=="暂停":self._pause_btn=btn
            if t=="继续":self._resume_btn=btn
        # Right-click pause = immediate hard pause
        self._pause_btn.bind("<Button-3>",lambda e:self._hard_pause())
        # Disable resume button initially
        self._resume_btn.config(state=tk.DISABLED)
        self._graceful_pause=False
        self.pb=ttk.Progressbar(t1,mode="indeterminate");self.pb.pack(fill=tk.X,padx=10,pady=2)
        self.sl=tk.Label(t1,text="就绪",font=("Arial",9),bg=self.BG,fg="#78909C");self.sl.pack()
        tk.Label(t1,text="采集结果",font=("Arial",10,"bold"),bg=self.BG,fg=self.T).pack(anchor=tk.W,padx=10)
        cols=("name","color","price","discount","sizes","imgs","tags","desc_imgs")
        self.tv=ttk.Treeview(t1,columns=cols,show="headings",height=8)
        for cid,ch,cw in zip(cols,["商品名称","颜色","价格","折扣","尺码","图片","产品标签","详情图"],[320,55,55,40,140,40,110,55]):
            self.tv.heading(cid,text=ch);self.tv.column(cid,width=cw)
        sb=tk.Scrollbar(t1,orient=tk.VERTICAL,command=self.tv.yview)
        self.tv.configure(yscrollcommand=sb.set)
        self.tv.pack(fill=tk.BOTH,expand=True,padx=(10,0),pady=2,side=tk.LEFT)
        sb.pack(fill=tk.Y,padx=(0,10),pady=2,side=tk.RIGHT)
        # Right-click context menu
        self._ctx_menu=tk.Menu(self.tv,tearoff=0)
        self._ctx_menu.add_command(label="删除选中行",command=self._del_row)
        self.tv.bind("<Button-3>" if sys.platform=="win32" else "<Button-2>",self._on_right_click)

        # Tab 2: Fetch Links
        t2=tk.Frame(nb,bg=self.BG);nb.add(t2,text="  抓取链接  ")
        tk.Label(t2,text="输入店铺/搜索页URL（每行一个），自动提取全部商品链接:",font=("Arial",10,"bold"),bg=self.BG,fg=self.T).pack(anchor=tk.W,padx=10,pady=(8,2))
        f1=tk.Frame(t2,bg=self.BG);f1.pack(fill=tk.X,padx=10,pady=5)
        self.fe=tk.Text(f1,height=3,font=("Consolas",10),relief="solid",bd=1,bg=self.W);self.fe.pack(fill=tk.X,padx=(0,5))
        self._fetch_btn=tk.Button(f1,text="抓取链接",font=("Arial",9),bg="#1565C0",fg=self.W,padx=10,pady=2,relief="flat",cursor="hand2",command=self._fl)
        self._fetch_btn.pack(side=tk.LEFT,padx=2)
        self._stop_fetch_btn=tk.Button(f1,text="停止抓取",font=("Arial",9),bg="#C62828",fg=self.W,padx=10,pady=2,relief="flat",cursor="hand2",command=self._stop_fl,state=tk.DISABLED)
        self._stop_fetch_btn.pack(side=tk.LEFT,padx=2)
        tk.Button(f1,text="追加到采集框",font=("Arial",9),bg="#EF6C00",fg=self.W,padx=10,pady=2,relief="flat",cursor="hand2",command=self._fa).pack(side=tk.LEFT,padx=2)
        tk.Button(f1,text="清空URL",font=("Arial",9),bg="#78909C",fg=self.W,padx=10,pady=2,relief="flat",cursor="hand2",command=self._clear_fe).pack(side=tk.LEFT,padx=2)
        tk.Label(t2,text="提取结果:",font=("Arial",10,"bold"),bg=self.BG,fg=self.T).pack(anchor=tk.W,padx=10,pady=(8,2))
        self.lt=scrolledtext.ScrolledText(t2,height=14,font=("Consolas",10),relief="solid",bd=1,bg="#FFF8E1")
        self.lt.pack(fill=tk.BOTH,expand=True,padx=10,pady=2)
        tk.Button(t2,text="全选复制到剪贴板",font=("Arial",9),bg="#FF6F00",fg=self.W,padx=10,pady=3,relief="flat",
                  cursor="hand2",command=lambda:self._cp()).pack()
        tk.Label(t2,text="提示: 1688 HTTP直抓; Shein 打开Chrome自动滚动加载",font=("Arial",8),bg=self.BG,fg="#78909C").pack(anchor=tk.W,padx=10,pady=3)
        # --- Manual browser mode ---
        mf=tk.Frame(t2,bg=self.BG);mf.pack(fill=tk.X,padx=10,pady=(10,2))
        tk.Label(mf,text="手动模式:",font=("Arial",10,"bold"),bg=self.BG,fg=self.T).pack(side=tk.LEFT,padx=(0,8))
        tk.Button(mf,text="打开浏览器",font=("Arial",9),bg="#2E7D32",fg=self.W,padx=8,pady=2,relief="flat",cursor="hand2",
                  command=self._open_manual_browser).pack(side=tk.LEFT,padx=2)
        tk.Button(mf,text="抓取本页链接",font=("Arial",9),bg="#1565C0",fg=self.W,padx=8,pady=2,relief="flat",cursor="hand2",
                  command=self._fetch_current_page_links).pack(side=tk.LEFT,padx=2)
        tk.Button(mf,text="关闭浏览器",font=("Arial",9),bg="#C62828",fg=self.W,padx=8,pady=2,relief="flat",cursor="hand2",
                  command=self._close_manual_browser).pack(side=tk.LEFT,padx=2)
        tk.Label(mf,textvariable=self._manual_status,font=("Arial",9),bg=self.BG,fg="#FF8F00").pack(side=tk.LEFT,padx=(8,0))
        tk.Label(t2,text="手动翻页→点「抓取本页链接」→切换页面→再点抓取，浏览器不会关",font=("Arial",8),bg=self.BG,fg="#78909C").pack(anchor=tk.W,padx=10)

        # === Shared Log ===
        lf=tk.Frame(pw,bg=self.BG)
        tk.Label(lf,text="运行日志",font=("Arial",9,"bold"),bg=self.BG,fg=self.T).pack(anchor=tk.W)
        self.lg=scrolledtext.ScrolledText(lf,font=("Consolas",9),bg="#ECEFF1",fg="#37474F",relief="solid",bd=1)
        self.lg.pack(fill=tk.BOTH,expand=True)
        pw.add(lf,minsize=100)
        self.products=[];self._urls=[];self._url_idx=0
        self.rt.after(200,self._load_session)

    def _lm(self,m):
        ts=datetime.now().strftime('%H:%M:%S')
        line=f"[{ts}] {m}"
        self.lg.insert(tk.END,line+"\n");self.lg.see(tk.END);self.rt.update_idletasks()
        # Write to log file
        try:
            with open(_LOG_FILE,'a',encoding='utf-8') as f:
                f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {m}\n")
        except:pass
    def _su(self):
        try:t=self.ut.get("1.0",tk.END).strip();open(self.UF,'w',encoding='utf-8').write(t)
        except:pass
        self._update_url_count()
    def _update_url_count(self):
        try:
            t=self.ut.get("1.0",tk.END).strip()
            n=len([l for l in t.split("\n")if l.strip()])
            self.ut_label.config(text=f"{n} URLs")
        except:pass
    def _oc(self):
        self._su();self._save_session()
        try:
            if self._manual_page:self._manual_page.close()
            if self._manual_ctx:self._manual_ctx.close()
            if self._manual_pw:self._manual_pw.stop()
        except:pass
        self.rt.destroy()
    def _ds(self):
        try:
            s=self.ut.tag_ranges(tk.SEL)
            if s:self.ut.delete(s[0],s[1])
            else:self.ut.delete("insert linestart","insert lineend+1c")
            self._su()
        except:pass
    def _dd(self):
        t=self.ut.get("1.0",tk.END);ls=[l.strip()for l in t.split("\n")if l.strip()];seen=set();ul=[]
        for l in ls:
            if l not in seen:seen.add(l);ul.append(l)
        self.ut.delete("1.0",tk.END);self.ut.insert("1.0","\n".join(ul));self._su()
    def _cu(self):self.ut.delete("1.0",tk.END);self._su()
    def _ca(self):
        t=self.ut.get("1.0",tk.END).strip()
        if t:self.rt.clipboard_clear();self.rt.clipboard_append(t);self._lm("Copied")
    def _cp(self):
        t=self.lt.get("1.0",tk.END).strip()
        if t:self.rt.clipboard_clear();self.rt.clipboard_append(t);self._lm("Links copied")
    def _cl(self):self.tv.delete(*self.tv.get_children());self.products.clear();self._save_session()
    def _on_right_click(self,event):
        item=self.tv.identify_row(event.y)
        if item:
            self.tv.selection_set(item)
            self._ctx_menu.post(event.x_root,event.y_root)
    def _del_row(self):
        sel=self.tv.selection()
        for item in sel:
            idx=self.tv.index(item)
            self.tv.delete(item)
            if idx<len(self.products):del self.products[idx]
        self._save_session()

    def _save_session(self):
        try:
            rows=[]
            for item in self.tv.get_children():
                vals=self.tv.item(item)['values']
                if len(vals)>=8:
                    rows.append({"name":vals[0]or"","color":vals[1]or"","price":vals[2]or"",
                                 "discount":vals[3]or"","sizes":vals[4]or"","imgs":vals[5]or"",
                                 "tags":vals[6]or"","desc_imgs":vals[7]or""})
            data={"products":self.products[:],"tv_rows":rows,
                  "url_idx":self._url_idx if hasattr(self,'_url_idx')else 0,
                  "platform":self.pl.get(),"region":self.sr.get(),
                  "log":self.lg.get("1.0",tk.END).strip(),
                  "_win_geo":self.rt.geometry()}
            with open(self.SESSION_FILE,'w',encoding='utf-8')as f:
                json.dump(data,f,ensure_ascii=False,indent=2)
        except Exception as e:
            self._lm(f"Session save failed: {e}")

    def _load_session(self):
        if not os.path.exists(self.SESSION_FILE):return False
        try:
            with open(self.SESSION_FILE,'r',encoding='utf-8')as f:
                data=json.load(f)
            self.products=data.get("products",[])
            cols=["name","color","price","discount","sizes","imgs","tags","desc_imgs"]
            for r in data.get("tv_rows",[]):
                self.tv.insert("","end",values=[r.get(k,"")for k in cols])
            saved_idx=data.get("url_idx",0)
            self._saved_url_idx=saved_idx
            if data.get("platform"):self.pl.set(data["platform"])
            if data.get("region"):self.sr.set(data["region"])
            if data.get("log"):
                self.lg.insert(tk.END,data["log"])
            self._lm(f"--- 自动恢复 {len(self.products)} 件商品记录 (上次采集到第 {saved_idx} 个) ---")
            return True
        except Exception as e:
            self._lm(f"Session restore failed: {e}")
            return False

    def _restore_session(self):
        if os.path.exists(self.SESSION_FILE):
            self._load_session()
        if not self.products:
            messagebox.showinfo("提示","没有可恢复的采集进度")
            return
        idx=getattr(self,'_saved_url_idx',0)
        # Read current URLs from input box
        t=self.ut.get("1.0",tk.END).strip()
        urls=[l.strip()for l in t.split("\n")if l.strip()]
        total=len(urls)
        # Enable resume button
        self._resume_btn.config(state=tk.NORMAL)
        self._pause_btn.config(state=tk.DISABLED)
        self._lm(f"已恢复 {len(self.products)} 件商品, 上次采集到第 {idx}/{total} 个URL")
        if idx<total:
            self.sl.config(text=f"点击「继续」从第 {idx+1}/{total} 个URL继续采集")
        else:
            self.sl.config(text=f"已恢复 {len(self.products)} 件 | {total} 个URL已全部采集完")

    def _fl(self):
        t=self.fe.get("1.0",tk.END).strip()
        if not t:return
        urls=[l.strip()for l in t.split("\n")if l.strip()]
        self._fetch_stop=False
        self._fetch_btn.configure(state=tk.DISABLED)
        self._stop_fetch_btn.configure(state=tk.NORMAL)
        threading.Thread(target=self._do_fl,args=(urls,),daemon=True).start()
    def _stop_fl(self):
        self._fetch_stop=True
        self._lm("Stopping...")
    def _clear_fe(self):
        self.fe.delete("1.0",tk.END)
    def _do_fl(self,urls):
        try:
            plat=self.pl.get()
            if plat=="1688":
                all_links=[]
                for url in urls:
                    if self._fetch_stop:break
                    try:
                        import requests
                        self._lm(f"Fetching: {url[:80]}...")
                        h={'User-Agent':'Mozilla/5.0'}
                        r=requests.get(url,headers=h,timeout=20)
                        links=list(set(re.findall(r'https?://detail\.1688\.com/offer/\d+\.html',r.text)))
                        all_links.extend(links)
                        self._lm(f"  Got {len(links)} links")
                    except Exception as e:self._lm(f"  Failed: {e}")
                self.lt.delete("1.0",tk.END)
                if all_links:self.lt.insert("1.0","\n".join(all_links));self._lm(f"Total: {len(all_links)} links")
                else:self._lm("No links found")
                self._fetch_done()
                return
            # AliExpress: browser with proxy, scroll + load more
            if plat=="aliexpress":
                from patchright.sync_api import sync_playwright
                import os as _os
                proxy_server = _os.getenv("PROXY_SERVER","")
                self._lm(f"Starting browser for {len(urls)} AliExpress URL(s)...")
                pw=sync_playwright().start()
                ctx_opts = {
                    "user_data_dir": _CHROME_PROFILE,
                    "executable_path": _CLOAKBROWSER if _CLOAKBROWSER else None,
                    "headless": False,
                    "args": ["--no-sandbox"],
                    "ignore_default_args": ["--enable-automation","--enable-unsafe-swiftshader"],
                    "viewport": {"width":1920,"height":1080},
                }
                if proxy_server:
                    ctx_opts["proxy"] = {"server": proxy_server}
                    self._lm(f"  Using proxy: {proxy_server}")
                ctx=pw.chromium.launch_persistent_context(**ctx_opts)
                page=ctx.new_page()
                all_links=[];prev_page_links=set()

                def _extract_links():
                    return page.evaluate("Array.from(document.querySelectorAll('a[href*=\"/item/\"]')).map(a => a.href).filter(h => /\\/item\\/\\d+\\.html/.test(h)).filter((h,idx,a) => a.indexOf(h) === idx)")

                for s_url in urls:
                    if self._fetch_stop:break
                    page.goto(s_url, timeout=90000, wait_until="networkidle")
                    # Wait for product links to render
                    try:
                        page.wait_for_selector('a[href*="/item/"]', timeout=20000)
                    except:
                        self._lm("  Products didn't load");break
                    time.sleep(0.5)

                    # Detect page type: wholesale (/w/wholesale-*) or store (/store/*)
                    is_wholesale = '/w/wholesale-' in s_url
                    is_store = '/store/' in s_url

                    if is_wholesale:
                        # === Wholesale: URL-based pagination (page=1, page=2, ...) ===
                        pg_num = 1
                        while True:
                            if self._fetch_stop:break
                            self._lm(f"  Page {pg_num}...")
                            links = _extract_links()
                            new = [l for l in links if l.split('?')[0] not in prev_page_links]
                            if not new:
                                self._lm(f"  No new links, done");break
                            all_links.extend(new)
                            prev_page_links.update(l.split('?')[0] for l in new)
                            self._lm(f"  Got {len(new)} new (total {len(all_links)})")
                            # Update textbox in real-time
                            self.lt.delete("1.0",tk.END)
                            self.lt.insert("1.0","\n".join(all_links))
                            # Build next page URL
                            pg_num += 1
                            if '?' in s_url:
                                # Check if page param already exists
                                import re as _re
                                if 'page=' in s_url:
                                    next_url = _re.sub(r'page=\d+', f'page={pg_num}', s_url)
                                else:
                                    next_url = s_url + f'&page={pg_num}'
                            else:
                                next_url = s_url + f'?page={pg_num}'
                            self._lm(f"  Loading page {pg_num}...")
                            page.goto(next_url, timeout=60000, wait_until="networkidle")
                            try:
                                page.wait_for_selector('a[href*="/item/"]', timeout=15000)
                            except:
                                self._lm(f"  Page {pg_num} load timeout, stopping");break
                            time.sleep(2)
                    else:
                        # === Store page: click page numbers in [currentpage] ===
                        pg_num = 1
                        while True:
                            if self._fetch_stop:break
                            self._lm(f"  Page {pg_num}...")
                            links = _extract_links()
                            new = [l for l in links if l.split('?')[0] not in prev_page_links]
                            if not new:
                                self._lm(f"  No new links");break
                            all_links.extend(new)
                            prev_page_links.update(l.split('?')[0] for l in new)
                            self._lm(f"  Got {len(new)} new (total {len(all_links)})")
                            # Update textbox in real-time
                            self.lt.delete("1.0",tk.END)
                            self.lt.insert("1.0","\n".join(all_links))
                            # Get current/total page from [currentpage] attribute
                            cur_page = page.evaluate("var c = document.querySelector('[currentpage]'); c ? parseInt(c.getAttribute('currentpage')) : 0;")
                            total_page = page.evaluate("var c = document.querySelector('[currentpage]'); c ? parseInt(c.getAttribute('totalpage')) : 0;")
                            if cur_page >= total_page:
                                self._lm(f"  Reached last page ({cur_page}/{total_page})");break
                            # Scroll pagination into view
                            page.evaluate("var c = document.querySelector('[currentpage]'); if(c) c.scrollIntoView({block:'center'});")
                            time.sleep(1)
                            # Click the next arrow (last child of [currentpage])
                            next_btn = page.locator('[currentpage] > :last-child')
                            if next_btn.count():
                                next_btn.click()
                            else:
                                self._lm("  No pagination button");break
                            # Poll for page change
                            changed = False
                            for _ in range(15):
                                time.sleep(2)
                                pg_now = page.evaluate("var c = document.querySelector('[currentpage]'); c ? parseInt(c.getAttribute('currentpage')) : 0;")
                                if pg_now != cur_page:
                                    changed = True;break
                            if not changed:
                                self._lm("  Retrying...")
                                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                                time.sleep(0.5)
                                page.evaluate("() => { var c = document.querySelector('[currentpage]'); if (c && c.children.length) { var last = c.children[c.children.length - 1]; var a = last.querySelector('a'); if (a) a.click(); } }")
                                time.sleep(3)
                            pg_num += 1

                page.close();ctx.close();pw.stop()
                if all_links:
                    seen=set();deduped=[]
                    for l in all_links:
                        base=l.split('?')[0]
                        if base not in seen:seen.add(base);deduped.append(l)
                    self.lt.delete("1.0",tk.END)
                    self.lt.insert("1.0","\n".join(deduped))
                    self._lm(f"Total: {len(deduped)} links from {len(urls)} pages")
                else:self._lm("No links found")
                self._fetch_done()
                return
            # Shein: use shared browser (same as main scraper)
            from patchright.sync_api import sync_playwright
            self._lm(f"Starting browser for {len(urls)} store URL(s)...")
            pw=sync_playwright().start()
            ctx=pw.chromium.launch_persistent_context(
                user_data_dir=_CHROME_PROFILE,
                executable_path=_CLOAKBROWSER if _CLOAKBROWSER else None,
                headless=False,
                args=["--no-sandbox"],
                ignore_default_args=["--enable-automation","--enable-unsafe-swiftshader"],
                viewport={"width":1920,"height":1080})
            page=ctx.new_page()
            all_links=[]
            for i,url in enumerate(urls,1):
                if self._fetch_stop:
                    self._lm("Stopped by user")
                    break
                self._lm(f"[{i}/{len(urls)}] Opening: {url[:100]}...")
                try:
                    page.goto(url, timeout=60000, wait_until="domcontentloaded")
                    time.sleep(3)
                    # CAPTCHA check (Shein: URL-based detection)
                    cw=0
                    while cw<60:
                        cu=page.url
                        if "challenge" in cu or "risk" in cu:
                            if cw==0:self._lm("  [CAPTCHA] Waiting...")
                            time.sleep(3);cw+=3
                        else:
                            break
                    # Scroll to load all products
                    prev=0;stable=0
                    for _ in range(80):
                        try:
                            page.evaluate("""
                                window.scrollTo(0, document.body.scrollHeight);
                                document.querySelectorAll('*').forEach(el => {
                                    if (el.scrollHeight > el.clientHeight) el.scrollTop = el.scrollHeight;
                                });
                            """)
                        except:pass
                        time.sleep(0.8)
                        try:cur=page.evaluate("document.querySelectorAll('a[href*=\"-p-\"]').length")
                        except:cur=prev
                        try:
                            for sel in ["button:has-text('Show more')","button:has-text('Load more')",
                                        "button:has-text('더 보기')","button:has-text('查看更多')",
                                        "[data-testid='load-more']",".load-more",".show-more"]:
                                btn=page.query_selector(sel)
                                if btn and btn.is_visible():btn.click();time.sleep(1)
                        except:pass
                        if cur==prev:
                            stable+=1
                            if stable>=3:break
                        else:stable=0
                        prev=cur
                    # Extract
                    links=page.evaluate("""
                        Array.from(document.querySelectorAll('a[href*="-p-"]'))
                            .map(a => a.href)
                            .filter(h => /-p-\\d+\\.html/.test(h))
                            .filter((h,idx,a) => a.indexOf(h) === idx)
                    """)
                    all_links.extend(links)
                    self._lm(f"  Got {len(links)} links (total {len(all_links)})")
                except Exception as e:
                    self._lm(f"  Failed: {e}")
            page.close();ctx.close();pw.stop()
            if all_links:
                seen=set();deduped=[]
                for l in all_links:
                    if l not in seen:seen.add(l);deduped.append(l)
                self.lt.delete("1.0",tk.END)
                self.lt.insert("1.0","\n".join(deduped))
                self._lm(f"Total: {len(deduped)} links from {len(urls)} stores")
            else:self._lm("No links found")
        finally:
            self._fetch_done()
    def _fetch_done(self):
        self._fetch_btn.configure(state=tk.NORMAL)
        self._stop_fetch_btn.configure(state=tk.DISABLED)
    # === Manual browser mode ===
    def _open_manual_browser(self):
        if self._manual_page:
            self._lm("Browser already running");return
        t=self.fe.get("1.0",tk.END).strip()
        urls=[l.strip() for l in t.split("\n") if l.strip()] if t else []
        t_url=urls[0] if urls else None
        self._lm("Launching manual browser...")
        threading.Thread(target=self._do_open_manual, args=(t_url,), daemon=True).start()

    def _do_open_manual(self, start_url):
        try:
            from patchright.sync_api import sync_playwright
            import os as _os
            proxy_server=_os.getenv("PROXY_SERVER","")
            pw=sync_playwright().start()
            ctx_opts={
                "user_data_dir":_CHROME_PROFILE,
                "executable_path":_CLOAKBROWSER if _CLOAKBROWSER else None,
                "headless":False,
                "args":["--no-sandbox"],
                "ignore_default_args":["--enable-automation","--enable-unsafe-swiftshader"],
                "viewport":{"width":1920,"height":1080},
            }
            if proxy_server:
                ctx_opts["proxy"]={"server":proxy_server}
            ctx=pw.chromium.launch_persistent_context(**ctx_opts)
            page=ctx.new_page()
            if start_url:
                page.goto(start_url, timeout=30000, wait_until="domcontentloaded")
            time.sleep(1)  # let initial render finish
            self._manual_pw=pw;self._manual_ctx=ctx;self._manual_page=page
            self._manual_status.set("浏览器已启动")
            self._lm("Manual browser ready — navigate & filter, then click 「抓取本页链接」")
        except Exception as e:
            self._lm(f"Failed to open browser: {e}")
            self._manual_status.set("启动失败")

    def _fetch_current_page_links(self):
        if not self._manual_page:
            self._lm("No browser running. Click 「打开浏览器」 first.");return
        try:
            url=self._manual_page.url
            self._lm(f"Fetching: {url[:120]}...")
            links=self._manual_page.evaluate(
                "Array.from(document.querySelectorAll('a[href*=\"/item/\"]')).map(a => a.href).filter(h => /\\/item\\/\\d+\\.html/.test(h)).filter((h,idx,a) => a.indexOf(h) === idx)")
            if not links:
                links=self._manual_page.evaluate(
                    "Array.from(document.querySelectorAll('a[href*=\"-p-\"]')).map(a => a.href).filter(h => /-p-\\d+\\.html/.test(h)).filter((h,idx,a) => a.indexOf(h) === idx)")
            if not links:
                self._lm("  No product links on this page");return
            existing=set()
            try:
                t=self.lt.get("1.0",tk.END).strip()
                existing=set(l.strip() for l in t.split("\n") if l.strip())
            except:pass
            new=[l for l in links if l.split('?')[0] not in existing]
            if not new:
                self._lm(f"  All {len(links)} links already in list");return
            current=list(existing);current.extend(new)
            self.lt.delete("1.0",tk.END);self.lt.insert("1.0","\n".join(current))
            self._lm(f"  +{len(new)} links (total {len(current)})")
        except Exception as e:
            self._lm(f"  Fetch failed: {e}")

    def _close_manual_browser(self):
        if not self._manual_page:
            self._lm("No browser running");return
        try:
            self._manual_page.close();self._manual_ctx.close();self._manual_pw.stop()
        except:pass
        self._manual_pw=None;self._manual_ctx=None;self._manual_page=None
        self._manual_status.set("浏览器未启动")
        self._lm("Manual browser closed")

    def _fa(self):
        links=self.lt.get("1.0",tk.END).strip()
        if not links:return
        cur=self.ut.get("1.0",tk.END).strip()
        # Normalize Shein URLs: keep only -p-NNN.html for dedup
        def norm(u):
            u=re.sub(r'\?.*','',u.strip()).rstrip('/')
            m=re.search(r'(/[^/]*)(-p-\d+\.html)',u)
            if m:u=m.group(2)
            return u
        ex=set(norm(l)for l in cur.split("\n")if l.strip())if cur else set()
        new=[l.strip()for l in links.split("\n")if l.strip()and norm(l)not in ex]
        if new:
            if cur:self.ut.insert(tk.END,"\n")
            self.ut.insert(tk.END,"\n".join(new));self._su();self._lm(f"Added {len(new)} URLs")
        else:self._lm("No new URLs")

    def _start(self):
        t=self.ut.get("1.0",tk.END).strip()
        if not t:messagebox.showerror("Error","Enter URLs");return
        # Dedup URLs (normalize Shein: keep only -p-NNN.html + domain)
        raw_urls=[l.strip()for l in t.split("\n")if l.strip()]
        def norm(u):
            u=re.sub(r'\?.*','',u).rstrip('/')
            # For Shein: extract product ID as key
            m=re.search(r'(/[^/]*)(-p-\d+\.html)',u)
            if m:u=m.group(2)  # keep only -p-NNN.html as key
            return u
        seen=set();urls=[]
        for u in raw_urls:
            if norm(u)not in seen:seen.add(norm(u));urls.append(u)
        self._pause_event=threading.Event();self._pause_event.set()
        self._stop_event=threading.Event();self._skip_current=False
        self._graceful_pause=False
        self._urls=urls;self._url_idx=0
        self._pause_btn.config(state=tk.NORMAL)
        self._resume_btn.config(state=tk.DISABLED)
        self.sl.config(text="▶ 采集中...")
        threading.Thread(target=self._ds2,daemon=True).start()
    def _pause(self):
        if not self._pause_event.is_set():self._lm("Already paused");return
        self._graceful_pause=True
        self._lm("将在完成当前商品后暂停...")
        self.sl.config(text="⏸ 等待当前SKU完成...")
    def _hard_pause(self):
        self._graceful_pause=False
        self._pause_event.clear()
        self._lm("HARD PAUSE - 立即停止")
        self.sl.config(text="⏸ 已暂停(硬)")
        self._pause_btn.config(state=tk.DISABLED)
        self._resume_btn.config(state=tk.NORMAL)
    def _resume(self):
        self._graceful_pause=False
        # 1. Thread alive but paused: just unpause
        if hasattr(self,'_pause_event') and not self._pause_event.is_set():
            self._pause_event.set()
            self._lm("Resumed - 继续从上次链接采集")
            self.sl.config(text="▶ 采集中...")
            self._pause_btn.config(state=tk.NORMAL)
            self._resume_btn.config(state=tk.DISABLED)
            return
        # 2. Thread dead or first start: start new thread from saved position
        t=self.ut.get("1.0",tk.END).strip()
        if not t:messagebox.showerror("Error","URL框为空");return
        urls=[l.strip()for l in t.split("\n")if l.strip()]
        self._pause_event=threading.Event();self._pause_event.set()
        self._stop_event=threading.Event();self._skip_current=False
        self._graceful_pause=False
        self._urls=urls
        self._url_idx=getattr(self,'_saved_url_idx',0)
        self._lm(f"从第 {self._url_idx+1}/{len(urls)} 个URL继续采集")
        self._pause_btn.config(state=tk.NORMAL)
        self._resume_btn.config(state=tk.DISABLED)
        self.sl.config(text="▶ 采集中...")
        threading.Thread(target=self._ds2,daemon=True).start()
    def _skip_sku(self):
        self._skip_current=True;self._graceful_pause=False
        self._pause_event.set()
        self._lm("SKIPPING current URL...")
        self.sl.config(text="▶ 采集中...")
        self._pause_btn.config(state=tk.NORMAL)
        self._resume_btn.config(state=tk.DISABLED)
    def _end(self):
        self._stop_event.set();self._pause_event.set()
        self._graceful_pause=False
        import subprocess
        for _ in range(3):
            subprocess.run("taskkill /F /IM chrome.exe 2>nul",shell=True)
            subprocess.run("taskkill /F /IM chromium.exe 2>nul",shell=True)
            time.sleep(0.5)
        self._lm("Stopped! All browser processes killed.")
        self._pause_btn.config(state=tk.NORMAL)
        self._resume_btn.config(state=tk.DISABLED)
    def _li(self):threading.Thread(target=self._dl,daemon=True).start()
    def _dl(self):
        from patchright.sync_api import sync_playwright
        plat=self.pl.get()
        if plat=="shein":
            lu="https://us.shein.com"if self.sr.get()=="US"else"https://kr.shein.com"
        elif plat=="1688":lu="https://www.1688.com/"
        else:lu="https://www.aliexpress.com/"
        self._lm(f"Opening {plat} login...")
        try:
            pw=sync_playwright().start()
            ctx=pw.chromium.launch_persistent_context(
                user_data_dir=_CHROME_PROFILE,
                executable_path=_CLOAKBROWSER if _CLOAKBROWSER else None,
                headless=False,args=["--no-sandbox"],
                ignore_default_args=["--enable-automation","--enable-unsafe-swiftshader"])
            page=ctx.new_page();page.goto(lu,timeout=30000)
            self._lm(f"Browser open! Log into {plat}, then close window.")
            while True:
                try:page.title();time.sleep(1)
                except:break
            time.sleep(3);ctx.close();pw.stop();self._lm("Login saved.")
        except Exception as e:self._lm(f"ERROR: {e}")
    def _ds2(self):
        self.pb.start();plat=self.pl.get();self.sl.config(text=f"Scraping... ({plat})")
        # Load SKUs from persistent file + results table
        scraped_skus=set()
        if os.path.exists(self.SKU_FILE):
            with open(self.SKU_FILE,'r',encoding='utf-8')as f:
                for line in f:
                    s=line.strip()
                    if s:scraped_skus.add(s)
        for p in self.products:
            sku=p.get("sku","")
            if sku:scraped_skus.add(sku)
        # Create ONE shared browser + ONE shared page for the entire session
        shared_page=None
        _shared_pw=None
        _shared_ctx=None
        try:
            from patchright.sync_api import sync_playwright as _sp2
            _shared_pw=_sp2().start()
            import random as _rnd
            _shared_ctx=_shared_pw.chromium.launch_persistent_context(
                user_data_dir=_CHROME_PROFILE,
                executable_path=_CLOAKBROWSER if _CLOAKBROWSER else None,
                headless=False,
                args=[f"--fingerprint={_rnd.randint(10000,99999)}"],
                ignore_default_args=[
                    "--enable-automation","--enable-unsafe-swiftshader",
                    "--disable-field-trial-config","--disable-breakpad",
                    "--disable-dev-shm-usage","--no-default-browser-check",
                    "--disable-background-networking","--disable-background-timer-throttling",
                    "--disable-backgrounding-occluded-windows",
                ],
                viewport={"width":1920,"height":1080})
            shared_page=_shared_ctx.new_page()
            shared_page.goto("about:blank")
            shared_page.evaluate("window.blur()")
            _page_count=0
            self._lm("Browser ready (single shared page — no focus steal)")
        except Exception as _e:
            self._lm(f"Failed to launch shared browser: {_e}")
            shared_page=None
        try:
            while self._url_idx<len(self._urls)and not self._stop_event.is_set():
                self._pause_event.wait()
                if self._stop_event.is_set():break
                url=self._urls[self._url_idx]
                self._url_idx+=1
                self._lm(f"[{self._url_idx}/{len(self._urls)}] {url[:80]}...")
                try:
                    if plat=="shein":
                        def on_captcha():
                            self.rt.after(0,lambda:messagebox.showinfo("人机验证","请在浏览器窗口完成验证。\n完成后将自动继续采集。"))
                        skip_fn=lambda:self._skip_current
                        prods=extract_shein(url,self._pause_event,scraped_skus,on_captcha,shared_page,skip_fn)
                        if prods is None:
                            self._lm(f"  >>> SKU DUPLICATE, skip <<<")
                            continue
                    else:
                        # AliExpress: check URL-level SKU dedup before scraping
                        if plat=="aliexpress":
                            ae_sku=re.search(r'/item/(\d+)\.html',url)
                            if not ae_sku:ae_sku=re.search(r'(\d{10,})',url)
                            if ae_sku:
                                ae_sku_id=ae_sku.group(1)
                                if ae_sku_id in scraped_skus:
                                    self._lm(f"  [SKIP URL] {ae_sku_id} already collected")
                                    continue
                        prods=extract_1688(url,shared_page,self._pause_event)if plat=="1688"else extract_aliexpress(url,shared_page,self._pause_event)
                    if prods:
                        for p in prods:
                            sku=p.get("sku","")
                            if sku and sku in scraped_skus:
                                self._lm(f"  [SKIP SKU] {sku} already collected")
                                continue
                            if sku:
                                scraped_skus.add(sku)
                                with open(self.SKU_FILE,'a',encoding='utf-8')as f:
                                    f.write(sku+'\n')
                            self.products.append(p)
                            sz=", ".join(p.get("sizes",[])[:8])
                            self.tv.insert("",tk.END,values=(
                                p.get("name","")[:40],
                                p.get("color",""),
                                p.get("price",""),
                                p.get("discount",""),
                                sz,
                                len(p.get("images",[])),
                                (p.get("product_tag","") or ", ".join(p.get("tags",[])[:2]))[:30],
                                len(p.get("description_images",[])),
                            ))
                            self._lm(f"  [{p.get('color','?')}] {p.get('price','?')} | {len(p.get('sizes',[]))} sizes | {len(p.get('images',[]))} imgs | {len(p.get('description_images',[]))} desc")
                            # Force GC every 5 products to reclaim HTML/BS memory
                            if _page_count % 5 == 0:
                                import gc; gc.collect()
                    else:self._lm("  FAIL: no data")
                except Exception as e:self._lm(f"  ERROR: {e}")
                finally:
                    self._skip_current=False
                    _page_count+=1
                    # Refresh page every 50 products to prevent Chrome slowdown
                    if shared_page and _page_count % 50 == 0:
                        try:
                            shared_page.close()
                            shared_page=_shared_ctx.new_page()
                            shared_page.goto("about:blank")
                            shared_page.evaluate("window.blur()")
                            self._lm(f"  [Page refreshed at {_page_count}]")
                        except:pass
                    # Random 3-5s delay between every URL (including skips)
                    import random
                    delay=random.uniform(3,5)
                    self._lm(f"  [wait {delay:.1f}s]")
                    time.sleep(delay)
                    # Graceful pause: stop after current URL completes
                    if self._graceful_pause:
                        self._pause_event.clear()
                        self._graceful_pause=False
                        self.sl.config(text=f"⏸ 已暂停 | 剩余 {len(self._urls)-self._url_idx} 个URL")
                        self._pause_btn.config(state=tk.DISABLED)
                        self._resume_btn.config(state=tk.NORMAL)
                        self._save_session()
                        self._pause_event.wait()
        finally:
            if shared_page:
                try:shared_page.close()
                except:pass
            if _shared_ctx:
                try:_shared_ctx.close()
                except:pass
            if _shared_pw:
                try:_shared_pw.stop()
                except:pass
        self.pb.stop();self.sl.config(text=f"Done: {len(self.products)} products")
        self._save_session()
        if not self._stop_event.is_set():
            self.rt.after(0,lambda:messagebox.showinfo("Done",f"采集完成\n{len(self.products)} products"))
    def _ex(self):
        if not self.products:messagebox.showwarning("Tip","No data");return
        fp=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialdir=os.path.join(os.path.dirname(__file__),"Outputs"))
        if not fp:return
        try:
            from openpyxl import Workbook
            wb=Workbook();ws=wb.active;ws.title="Products"
            for c,h in enumerate(TEMPLATE_HEADERS,1):ws.cell(row=1,column=c,value=h)
            # Filter out entries with no product tag or no images
            filtered=[p for p in self.products if (p.get("product_tag","") or p.get("shein_category","")).strip() and p.get("images",[])]
            row=2
            for p in filtered:
                color=p.get("color","");imgs=p.get("images",[])
                desc_imgs=p.get("description_images",[])or[]
                tags=p.get("tags",[])or[]
                for size in p.get("sizes",["One Size"]):
                    ws.cell(row=row,column=1,value=p.get("parent_sku",""))
                    ws.cell(row=row,column=2,value=f"{p.get('parent_sku','')}{size}{color}")
                    ws.cell(row=row,column=3,value=p.get("name",""));ws.cell(row=row,column=4,value=p.get("description",""))
                    pt=p.get("product_tag","") or p.get("shein_category","")
                    ws.cell(row=row,column=5,value=pt if pt else ", ".join(tags[:3]))
                    ws.cell(row=row,column=6,value=p.get("sku",""))
                    if color:ws.cell(row=row,column=10,value=color)
                    ws.cell(row=row,column=11,value=size);ws.cell(row=row,column=12,value=p.get("url",""))
                    sp=p.get("size_prices",{});sz_price=sp.get(size,p.get("price",""))
                    ws.cell(row=row,column=13,value=sz_price);ws.cell(row=row,column=39,value=p.get("url",""))
                    ws.cell(row=row,column=55,value=p.get("name",""))
                    if imgs:ws.cell(row=row,column=18,value=imgs[0])
                    for j,img in enumerate(imgs[1:21]):ws.cell(row=row,column=19+j,value=img)
                    # Description images: append after main/sub images
                    for k, dimg in enumerate(desc_imgs[:20]):
                        ws.cell(row=row,column=67+k,value=dimg)
                    ci=p.get("color_images",{})
                    if color and color in ci:ws.cell(row=row,column=41,value=ci[color])
                    elif ci:ws.cell(row=row,column=41,value=list(ci.values())[0])
                    row+=1
            wb.save(fp);self._lm(f"Excel: {fp} ({row-2} rows, {len(self.products)-len(filtered)} filtered)");messagebox.showinfo("Done",f"Exported {row-2} rows\n{len(self.products)-len(filtered)} filtered (no tag/image)")
        except Exception as e:messagebox.showerror("Error",str(e))
    def run(self):self.rt.mainloop()

if __name__=="__main__":
    import updater
    root=tk.Tk();root.withdraw()
    if updater.check_and_update(root):
        root.destroy();App().run()
    else:
        root.destroy()
